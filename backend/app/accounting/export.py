"""Export services for accounting data (CSV, XLSX)."""


import csv
import io
from datetime import date

from sqlalchemy.ext.asyncio import AsyncSession

from app.accounting import service
from app.accounting.schemas import ExpenseFilter
from app.core.pagination import PaginationParams


async def export_expenses_csv(
    db: AsyncSession,
    filters: ExpenseFilter,
) -> bytes:
    """Export filtered expenses as a CSV file."""
    # Fetch all matching expenses (no pagination limit for export)
    pagination = PaginationParams(page=1, page_size=10000)
    expenses, _ = await service.list_expenses(db, filters, pagination)

    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow(
        [
            "Date",
            "Vendor",
            "Description",
            "Category",
            "Amount",
            "Currency",
            "Tax",
            "Payment Method",
            "Status",
            "Notes",
        ]
    )

    for expense in expenses:
        writer.writerow(
            [
                expense.date.isoformat() if isinstance(expense.date, date) else str(expense.date),
                expense.vendor_name or "",
                expense.description or "",
                expense.category.name if expense.category else "Uncategorized",
                f"{expense.amount:.2f}",
                expense.currency,
                f"{expense.tax_amount:.2f}" if expense.tax_amount else "",
                expense.payment_method.value if expense.payment_method else "",
                expense.status.value,
                expense.notes or "",
            ]
        )

    return output.getvalue().encode("utf-8")


async def export_expenses_xlsx(
    db: AsyncSession,
    filters: ExpenseFilter,
) -> bytes:
    """Export filtered expenses as an XLSX file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise RuntimeError(
            "openpyxl is required for XLSX export. Install it with: pip install openpyxl"
        )

    # Fetch all matching expenses
    pagination = PaginationParams(page=1, page_size=10000)
    expenses, _ = await service.list_expenses(db, filters, pagination)

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")

    headers = [
        "Date",
        "Vendor",
        "Description",
        "Category",
        "Amount",
        "Currency",
        "Tax",
        "Payment Method",
        "Status",
        "Notes",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, expense in enumerate(expenses, 2):
        ws.cell(
            row=row_idx,
            column=1,
            value=(
                expense.date.isoformat() if isinstance(expense.date, date) else str(expense.date)
            ),
        )
        ws.cell(row=row_idx, column=2, value=expense.vendor_name or "")
        ws.cell(row=row_idx, column=3, value=expense.description or "")
        ws.cell(
            row=row_idx,
            column=4,
            value=(expense.category.name if expense.category else "Uncategorized"),
        )
        ws.cell(row=row_idx, column=5, value=expense.amount)
        ws.cell(row=row_idx, column=6, value=expense.currency)
        ws.cell(row=row_idx, column=7, value=expense.tax_amount or "")
        ws.cell(
            row=row_idx,
            column=8,
            value=(expense.payment_method.value if expense.payment_method else ""),
        )
        ws.cell(row=row_idx, column=9, value=expense.status.value)
        ws.cell(row=row_idx, column=10, value=expense.notes or "")

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value else ""
            max_length = max(max_length, len(val))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
