"""Excel import logic for the cashbook module.

Parses the accountant's cashbook Excel template format:
- Cash_Book sheet: Date, Transaction, Bank Balance, Total Amount, then category columns
- Credit card sheet: Date, Description, Total, GST/HST, then category columns
"""

import io
from datetime import date, datetime

import openpyxl

from app.cashbook.models import EntryType
from app.cashbook.schemas import ImportPreview, ParsedExcelRow
from app.cashbook.service import calculate_tax


# Column header -> category name mapping for the Cash_Book sheet
CASHBOOK_INCOME_COLUMNS = {
    "grant": "Grant",
    "hst/gst collected": "HST/GST Collected",
    "fees": "Fees",
    "rental income": "Rental Income",
    "other": "Other Income",
}

CASHBOOK_EXPENSE_COLUMNS = {
    "hst/gst paid": "HST/GST Paid",
    "advertising": "Advertising",
    "inventory": "Inventory",
    "shipping": "Shipping",
    "fuel": "Fuel",
    "travel": "Travel",
    "credit card payment": "Credit Card Payment",
    "meals": "Meals",
    "depreciation expense": "Depreciation",
    "dues & subscriptions": "Dues & Subscriptions",
    "education & training": "Education & Training",
    "insurance – general liability": "Insurance General",
    "insurance – vehicles": "Insurance Vehicles",
    "interest expense": "Interest Expense",
    "meals and entertainment": "Meals & Entertainment",
    "office supplies": "Office Supplies",
    "professional fees": "Professional Fees",
    "rent expense": "Rent",
    "repairs & maintenance": "Repairs & Maintenance",
    "taxes – corporate tax": "Corporate Tax",
    "telephone – land line": "Telephone Land",
    "telephone – wireless": "Telephone Wireless",
    "travel expense": "Travel",
    "utilities": "Utilities",
    "vehicle – fuel": "Vehicle Fuel",
    "vehicle – repairs & maintenance": "Vehicle Repairs",
}

# Credit card sheet category columns
CREDIT_CARD_COLUMNS = {
    "meals": "Meals",
    "staff meals": "Meals",
    "petrol": "Vehicle Fuel",
    "office supplies": "Office Supplies",
    "office equipment": "Office Supplies",
    "travel": "Travel",
    "training": "Education & Training",
    "bank interest": "Interest Expense",
    "hotel - business conference": "Travel",
    "parking": "Travel",
    "bank fees": "Professional Fees",
    "team meeting": "Meals & Entertainment",
}


def _parse_date(value: object) -> date | None:
    """Try to parse a date from various formats."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        # Try common formats
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%B %d, %Y", "%b %d, %Y",
                     "%d %B %Y", "%d %b %Y", "%dth %B %Y", "%dst %B %Y",
                     "%dnd %B %Y", "%drd %B %Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        # Handle ordinal dates like "19th July 2023"
        import re
        cleaned = re.sub(r"(\d+)(st|nd|rd|th)", r"\1", value)
        for fmt in ("%d %B %Y", "%d %b %Y"):
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
    return None


def _parse_amount(value: object) -> float | None:
    """Parse a numeric amount."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value != 0 else None
    if isinstance(value, str):
        try:
            cleaned = value.replace("$", "").replace(",", "").strip()
            val = float(cleaned)
            return val if val != 0 else None
        except ValueError:
            return None
    return None


def parse_cashbook_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[ParsedExcelRow]:
    """Parse the Cash_Book sheet from the accountant's template."""
    rows: list[ParsedExcelRow] = []

    # Find the header row (look for "Date" and "Transaction" or "Total Amount")
    header_row = None
    for row_idx in range(1, min(ws.max_row + 1, 15)):
        for col_idx in range(1, min(ws.max_column + 1, 5)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val and isinstance(cell_val, str) and "date" in cell_val.lower().strip():
                header_row = row_idx
                break
        if header_row:
            break

    if header_row is None:
        return rows

    # Map column indices to category names
    col_map: dict[int, tuple[str, EntryType]] = {}
    date_col = None
    desc_col = None
    total_col = None

    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=header_row, column=col_idx).value
        if header is None:
            continue
        header_lower = str(header).strip().lower()

        if "date" in header_lower and date_col is None:
            date_col = col_idx
        elif header_lower in ("transaction", "description"):
            desc_col = col_idx
        elif header_lower == "total amount":
            total_col = col_idx
        else:
            # Check income columns
            for key, cat_name in CASHBOOK_INCOME_COLUMNS.items():
                if key in header_lower:
                    col_map[col_idx] = (cat_name, EntryType.INCOME)
                    break
            else:
                # Check expense columns
                for key, cat_name in CASHBOOK_EXPENSE_COLUMNS.items():
                    if key in header_lower:
                        col_map[col_idx] = (cat_name, EntryType.EXPENSE)
                        break

    if date_col is None or total_col is None:
        return rows

    # Parse data rows
    current_date: date | None = None
    for row_idx in range(header_row + 1, ws.max_row + 1):
        # Get date (may carry forward from previous row)
        raw_date = ws.cell(row=row_idx, column=date_col).value
        parsed_date = _parse_date(raw_date)
        if parsed_date:
            current_date = parsed_date

        # Get total amount
        raw_total = ws.cell(row=row_idx, column=total_col).value
        total = _parse_amount(raw_total)
        if total is None or total == 0:
            continue

        # Get description
        desc = ""
        if desc_col:
            raw_desc = ws.cell(row=row_idx, column=desc_col).value
            desc = str(raw_desc).strip() if raw_desc else ""

        # Determine category by finding non-zero category column
        category_name: str | None = None
        entry_type = EntryType.EXPENSE  # default
        for col_idx, (cat_name, cat_entry_type) in col_map.items():
            val = _parse_amount(ws.cell(row=row_idx, column=col_idx).value)
            if val and val != 0:
                category_name = cat_name
                entry_type = cat_entry_type
                break

        errors: list[str] = []
        if current_date is None:
            errors.append("Could not parse date")
        if not desc:
            errors.append("Missing description")

        rows.append(
            ParsedExcelRow(
                row_number=row_idx,
                sheet_name="Cash_Book",
                date=current_date,
                description=desc or f"Row {row_idx}",
                total_amount=abs(total),
                category_name=category_name,
                entry_type=entry_type,
                tax_amount=None,
                errors=errors,
            )
        )

    return rows


def parse_credit_card_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    default_tax_rate: float = 13.0,
) -> list[ParsedExcelRow]:
    """Parse the Credit card sheet from the accountant's template."""
    rows: list[ParsedExcelRow] = []

    # Find header row
    header_row = None
    for row_idx in range(1, min(ws.max_row + 1, 10)):
        for col_idx in range(1, min(ws.max_column + 1, 5)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val and isinstance(cell_val, str) and "date" in cell_val.lower().strip():
                header_row = row_idx
                break
        if header_row:
            break

    if header_row is None:
        return rows

    # Map columns
    date_col = None
    desc_col = None
    total_col = None
    gst_col = None
    category_cols: dict[int, str] = {}

    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=header_row, column=col_idx).value
        if header is None:
            continue
        header_lower = str(header).strip().lower()

        if "date" in header_lower and date_col is None:
            date_col = col_idx
        elif header_lower in ("description", "transaction"):
            desc_col = col_idx
        elif header_lower == "total":
            total_col = col_idx
        elif "gst" in header_lower or "hst" in header_lower:
            gst_col = col_idx
        else:
            for key, cat_name in CREDIT_CARD_COLUMNS.items():
                if key == header_lower:
                    category_cols[col_idx] = cat_name
                    break

    if total_col is None:
        return rows

    # Parse data rows
    current_date: date | None = None
    for row_idx in range(header_row + 1, ws.max_row + 1):
        # Skip totals row
        raw_date = ws.cell(row=row_idx, column=date_col).value if date_col else None
        parsed_date = _parse_date(raw_date)
        if parsed_date:
            current_date = parsed_date

        raw_total = ws.cell(row=row_idx, column=total_col).value
        total = _parse_amount(raw_total)
        if total is None or total == 0:
            continue

        # Check if this is a totals/summary row (last row)
        if row_idx >= ws.max_row:
            continue

        desc = ""
        if desc_col:
            raw_desc = ws.cell(row=row_idx, column=desc_col).value
            desc = str(raw_desc).strip() if raw_desc else ""

        # Get tax
        tax_amount: float | None = None
        if gst_col:
            tax_amount = _parse_amount(ws.cell(row=row_idx, column=gst_col).value)
        if tax_amount is None and default_tax_rate > 0:
            tax_amount = calculate_tax(abs(total), default_tax_rate)

        # Determine category
        category_name: str | None = None
        for col_idx, cat_name in category_cols.items():
            val = _parse_amount(ws.cell(row=row_idx, column=col_idx).value)
            if val and val != 0:
                category_name = cat_name
                break

        errors: list[str] = []
        if current_date is None:
            errors.append("Could not parse date")
        if not desc:
            errors.append("Missing description")

        rows.append(
            ParsedExcelRow(
                row_number=row_idx,
                sheet_name="Credit card",
                date=current_date,
                description=desc or f"Row {row_idx}",
                total_amount=abs(total),
                category_name=category_name,
                entry_type=EntryType.EXPENSE,
                tax_amount=round(tax_amount, 2) if tax_amount else None,
                errors=errors,
            )
        )

    return rows


def parse_excel_file(file_bytes: bytes, default_tax_rate: float = 13.0) -> ImportPreview:
    """Parse an uploaded Excel cashbook file and return a preview."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    all_rows: list[ParsedExcelRow] = []
    sheets_found: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_lower = sheet_name.lower()

        if "cash" in sheet_lower and "book" in sheet_lower:
            parsed = parse_cashbook_sheet(ws)
            if parsed:
                all_rows.extend(parsed)
                sheets_found.append(sheet_name)
        elif "credit" in sheet_lower:
            parsed = parse_credit_card_sheet(ws, default_tax_rate)
            if parsed:
                all_rows.extend(parsed)
                sheets_found.append(sheet_name)

    valid_rows = sum(1 for r in all_rows if not r.errors)
    error_rows = sum(1 for r in all_rows if r.errors)

    return ImportPreview(
        rows=all_rows,
        total_rows=len(all_rows),
        valid_rows=valid_rows,
        error_rows=error_rows,
        sheets_found=sheets_found,
    )
